import openpyxl

fileName = r"D:\Testing\PythonPractice\AutomationPractice\ExcelSheet.xlsx"
excelbook = openpyxl.load_workbook(fileName)
sheet = excelbook["Sheet1"]


print(sheet.cell(row=1, column=1).value)

excelbook.save(fileName)

rowCount = sheet.max_row
columCount = sheet.max_column

for i in range(1, rowCount+1):
    for j in range(1, columCount+1):
        print(sheet.cell(row=i, column=j).value)