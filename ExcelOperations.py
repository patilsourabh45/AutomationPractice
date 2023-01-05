### pip install openpyxl
## execute on cmd

## Install package in pycharm for your project

import openpyxl

filename = r"C:\tmp\Class\UserIDPass.xlsx"
# Locate and open /load file by file name
workbook = openpyxl.load_workbook(filename)

# Locate/ select perticuler sheet
###workbook.get_sheet_by_name('UserID')
sheet = workbook['UserID']

# read data from each cell
print(sheet.cell(row=1, column=1).value)
print(sheet.cell(row=1, column=2).value)

# "Manoj" "Manoj@123"
sheet.cell(row=10, column=1).value = "Shubham"
sheet.cell(row=10, column=2).value = "Shubham@123"

workbook.save(filename)

## assignment : Read all file data from each row


# import openpyxl
#
# filename = r"D:\Testing\PythonPractice\AutomationPractice\ExcelSheet.xlsx"
# excelBook = openpyxl.load_workbook(filename)
#
# sheet = excelBook["Sheet1"]
#
# print(sheet.cell(row=1, column=1).value)
# print(sheet.cell(row=1, column=2).value)
#
# sheet.cell (row=1, column=1).value = 1
# sheet.cell (row=1, column=2).value ="Sourabh"
#
# excelBook.save(filename)
#
# print(sheet.cell(row=1, column=1).value)
# print(sheet.cell(row=1, column=2).value)